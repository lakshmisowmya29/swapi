from openpyxl import load_workbook
import pymongo

def excel_to_json(fname):
    wb = load_workbook(fname)
    ws = wb.active
    mylist = []
    rcount = ws.max_row
    ccount = ws.max_column

    for row in range(1, rcount):
        mydict = {}
        for col in range(0, ccount):
            if ws[row][col].value is None:
                print (ws[row][col])
                mydict[ws[1][col].value] = ""
            else:
                mydict[ws[1][col].value] = ws[row+1][col].value
        mylist.append(mydict)
    print (mylist)
    return mylist

def store_menu_data_in_db(menu_data):
    myclient = pymongo.MongoClient("mongodb://localhost:27017/")
    mydb = myclient["auramenu"]
    mycol = mydb["wapimenu"]
    mydict = {"menu_data":menu_data}
    x = mycol.insert_one(mydict)
    print ("data is stored")
    return

def main():
    fname = 'aura_wapi_menu.xlsx'
    menu_data = excel_to_json(fname)
    #print (type(menu_data))
    #store_menu_data_in_db(menu_data)

if __name__ == "__main__":
    main()



